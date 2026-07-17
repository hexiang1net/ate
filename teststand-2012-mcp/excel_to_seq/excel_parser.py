"""Parse ATETestPlan Excel template into test case data structures."""

from typing import List, Tuple, Dict
from openpyxl import load_workbook

from .excel_to_seq_model import ParsedTestCase, ParsedViParameter, ParsedVariable


class ExcelParser:
    """Parse ATETestPlan_V1.xls Excel template."""

    # Column indices (1-based for openpyxl)
    # Sheet 1: ATE test item check list
    # Column order aligned with seq_to_excel (excel_generator.py) as the standard.
    COL_STEP_NO = 1          # A: Step_No
    COL_STEP = 2             # B: Step (startup/main/cleanup)
    COL_TEST_PROJECT = 3      # C: Test project items
    COL_DESCRIPTION = 4       # D: Description of Test Requirements
    COL_STEP_TYPE = 5        # E: step type (Action, NumericLimitTest, etc.)
    COL_INSTRUMENT_VI = 6     # F: instrument_vi (VI filename)
    COL_LIMITS = 7           # G: limits expression
    COL_USL = 8              # H: USL (upper spec limit)
    COL_LSL = 9             # I: LSL (lower spec limit)
    COL_UNIT = 10            # J: unit
    COL_FORMAT = 11          # K: format
    COL_RUN_MODE = 12        # L: run mode
    COL_WAIT_TIME = 13       # M: wait(s)
    COL_ADAPTER = 14         # N: adapter
    COL_ADDITIONAL_RESULTS = 15  # O: additional results
    COL_SETTINGS = 16        # P: PreExpr/StatusExpr/PostExpr settings
    COL_COMMENTS = 17        # Q: comments
    COL_INPUT_SIGNALS = 18    # R: Input Signals
    COL_OUTPUT_LOADS = 19    # S: Output Loads
    COL_PRECONDITION = 20     # T: Precondition Item/Status
    COL_COMMAND = 21          # U: VI path (Command - manual entry)
    COL_TEST_POINT = 22       # V: Test point
    COL_MEASUREMENT = 23     # W: Measurement Value
    COL_PRECISION = 24        # X: precision(%)
    COL_METHOD_CHECK = 25    # Y: Method Check
    COL_EQUIPMENT = 26       # Z: Equipment Solution
    COL_PARAMETER = 27        # AA: Parameter
    COL_LOGFILE = 28          # AB: Logfile confirm
    COL_TE_COMMENTS = 29     # AC: TE comments
    COL_REMARKS = 30          # AD: remarks

    def parse(self, excel_path: str) -> Tuple[List[ParsedTestCase], List[ParsedViParameter], List[ParsedVariable]]:
        """Parse Excel file.

        Returns:
            Tuple of (test_cases, vi_parameters, variables)
        """
        try:
            wb = load_workbook(excel_path, data_only=True)
        except Exception as e:
            err_msg = str(e)
            if "OLE2" in err_msg or "OOXML" in err_msg:
                raise ValueError(
                    f"文件不是有效的 .xlsx 格式: {excel_path}\n"
                    "请确认文件是真正的 .xlsx 格式（非 .xls 改名），或文件未损坏。"
                ) from e
            raise ValueError(f"无法读取 Excel 文件: {excel_path} — {err_msg}") from e

        # Parse Sheet 1: ATE test item check list
        ws1 = wb["ATE test item check list"]
        test_cases = self._parse_check_list_sheet(ws1)

        # Parse Sheet 2: vi_parameter_table
        ws2 = wb["vi_parameter_table"]
        vi_params = self._parse_vi_parameter_sheet(ws2)

        # Parse Sheet 3: Variables
        variables = []
        if "Variables" in wb.sheetnames:
            ws3 = wb["Variables"]
            variables = self._parse_variables_sheet(ws3)

        return test_cases, vi_params, variables

    def _parse_check_list_sheet(self, ws) -> List[ParsedTestCase]:
        """Parse Sheet 1 - data starts at row 5."""
        test_cases = []

        for row in ws.iter_rows(min_row=5, max_col=self.COL_REMARKS, values_only=True):
            step_no = row[self.COL_STEP_NO - 1]
            if not step_no:
                continue

            instrument_vi = str(row[self.COL_INSTRUMENT_VI - 1] or "").strip()
            vi_path = str(row[self.COL_COMMAND - 1] or "").strip()

            # Column S (Command) is usually empty; real VI paths come from
            # column E (instrument_vi).  Merge: prefer explicit vi_path,
            # fall back to instrument_vi.
            if not vi_path and instrument_vi:
                vi_path = instrument_vi

            tc = ParsedTestCase(
                step_no=str(step_no).strip(),
                step_group=self._normalize_step_group(row[self.COL_STEP - 1]),
                test_project=str(row[self.COL_TEST_PROJECT - 1] or "").strip(),
                step_type=str(row[self.COL_STEP_TYPE - 1] or "Action").strip(),
                instrument_vi=instrument_vi,
                limits=str(row[self.COL_LIMITS - 1] or "").strip(),
                unit=str(row[self.COL_UNIT - 1] or "").strip(),
                format=str(row[self.COL_FORMAT - 1] or "").strip(),
                run_mode=str(row[self.COL_RUN_MODE - 1] or "").strip(),
                wait_seconds=str(row[self.COL_WAIT_TIME - 1] or "").strip(),
                adapter=str(row[self.COL_ADAPTER - 1] or "").strip(),
                additional_results=str(row[self.COL_ADDITIONAL_RESULTS - 1] or "").strip(),
                settings=str(row[self.COL_SETTINGS - 1] or "").strip(),
                comments=str(row[self.COL_COMMENTS - 1] or "").strip(),
                description=str(row[self.COL_DESCRIPTION - 1] or "").strip(),
                vi_path=vi_path,
                usl=str(row[self.COL_USL - 1] or "").strip(),
                lsl=str(row[self.COL_LSL - 1] or "").strip(),
                method_check=str(row[self.COL_METHOD_CHECK - 1] or "").strip(),
                equipment=str(row[self.COL_EQUIPMENT - 1] or "").strip(),
            )
            test_cases.append(tc)

        return test_cases

    def _parse_vi_parameter_sheet(self, ws) -> List[ParsedViParameter]:
        """Parse Sheet 2 - VI parameter table.

        Parses VI definition rows and step rows to build a complete parameter map.
        Structure: VI definition row (no step_no, has vi_name) followed by step rows
        (has step_no, no vi_name) with parameter values.
        """
        vi_params = []
        current_vi_name = ""
        current_param_names = []
        step_params_map = {}  # step_no -> param_values

        for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
            step_no = row[0]
            vi_name = row[1]

            if step_no and vi_name:
                # New VI entry with both step_no and vi_name
                # Save previous VI's step params first
                if current_vi_name and step_params_map:
                    for sn, pvals in step_params_map.items():
                        vi_params.append(ParsedViParameter(
                            step_no=sn,
                            instrument_vi_name=current_vi_name,
                            param_names=list(current_param_names),
                            param_values=pvals,
                        ))
                # Start new VI
                current_vi_name = str(vi_name).strip() if vi_name else ""
                current_param_names = [str(row[i] or "").strip() for i in range(2, 12)]
                step_params_map = {}  # Reset for new VI
            elif not step_no and vi_name:
                # VI definition row - save previous VI entries first
                if current_vi_name and step_params_map:
                    for sn, pvals in step_params_map.items():
                        vi_params.append(ParsedViParameter(
                            step_no=sn,
                            instrument_vi_name=current_vi_name,
                            param_names=list(current_param_names),
                            param_values=pvals,
                        ))
                current_vi_name = str(vi_name).strip() if vi_name else ""
                current_param_names = [str(row[i] or "").strip() for i in range(2, 12)]
                step_params_map = {}
            elif step_no and not vi_name:
                # Step row - param values (no vi_name, but has step_no)
                current_step_no = str(step_no).strip() if step_no else ""
                param_values = [str(row[i] or "").strip() for i in range(2, 12)]
                step_params_map[current_step_no] = param_values

        # Save last VI entry
        if current_vi_name and step_params_map:
            for sn, pvals in step_params_map.items():
                vi_params.append(ParsedViParameter(
                    step_no=sn,
                    instrument_vi_name=current_vi_name,
                    param_names=list(current_param_names),
                    param_values=pvals,
                ))

        return vi_params

    def _parse_variables_sheet(self, ws) -> List[ParsedVariable]:
        """Parse Sheet 3 - Variables table.

        Columns: A=Container, B=Variable Name, C=Value, D=Type
        Data starts at row 2.
        """
        variables = []
        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            container = str(row[0] or "").strip()
            var_name = str(row[1] or "").strip()
            value = str(row[2] or "").strip()
            var_type = str(row[3] or "").strip()
            if container and var_name:
                variables.append(ParsedVariable(
                    container=container,
                    name=var_name,
                    value=value,
                    var_type=var_type,
                ))
        return variables

    def _normalize_step_group(self, value) -> str:
        """Normalize step group value."""
        if not value:
            return "main"
        v = str(value).lower().strip()
        if v in ("startup", "setup"):
            return "startup"
        if v in ("cleanup",):
            return "cleanup"
        return "main"


def group_by_parent_step(test_cases: List[ParsedTestCase]) -> Tuple[List[ParsedTestCase], Dict[str, List[ParsedTestCase]], Dict[str, str]]:
    """Group test cases by parent step number.

    For step_no like "2_1_1", the parent is "2_1".
    For step_no like "2_1", it is a top-level step (no parent).

    Returns:
        (top_level_cases, sub_sequence_groups, parent_names)
        where parent_names maps parent_no -> test_project name
    """
    groups = {}
    parent_names = {}
    top_level = []

    for tc in test_cases:
        parts = tc.step_no.split("_")
        if len(parts) >= 3:
            # Nested step (e.g., 2_1_1) -> parent is 2_1
            parent_no = "_".join(parts[:2])
            if parent_no not in groups:
                groups[parent_no] = []
                parent_names[parent_no] = tc.test_project
            groups[parent_no].append(tc)
        else:
            # Top-level step
            top_level.append(tc)

    return top_level, groups, parent_names


def parse_step_no(step_no: str) -> Tuple[str, int, int]:
    """Parse step number into components.

    Returns:
        (prefix, main_idx, sub_idx) - e.g., "2_1_3" -> ("2", 1, 3)
    """
    parts = step_no.split("_")
    if len(parts) == 2:
        return parts[0], int(parts[1]), 0
    elif len(parts) >= 3:
        return parts[0], int(parts[1]), int(parts[2])
    return "2", 0, 0