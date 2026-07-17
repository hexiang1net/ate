"""Excel generator for ATE test case report.

Generates Excel files matching the ATETestPlan_V1.xls template format.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from .testcase_model import TestCase, TestCaseReport, ViParameter


class ExcelGenerator:
    """Generate Excel test case report in ATETestPlan template format."""

    def __init__(self):
        self._setup_styles()

    def _setup_styles(self):
        """Define consistent styling matching template."""
        self.title_font = Font(bold=True, size=12)
        self.header_font = Font(bold=True, color="000000", size=10)
        self.header_fill_gray = PatternFill("solid", fgColor="D9D9D9")
        self.header_fill_yellow = PatternFill("solid", fgColor="FFFF00")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    @staticmethod
    def _write_cell_safe(ws, row, column, value):
        """Write a cell value, handling formula-like strings safely.
        Values starting with =, +, -, @ are written as inline strings to prevent
        Excel from interpreting them as formulas (which causes corruption errors).
        """
        cell = ws.cell(row=row, column=column)
        if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@'):
            cell._value = value
            cell.data_type = 's'
        else:
            cell.value = value
        return cell

    def generate(self, report: TestCaseReport, output_path: str) -> None:
        """Generate complete Excel report matching ATETestPlan_V1.xls template."""
        wb = Workbook()

        # Sheet 1: "ATE test item check list" - 4 rows template + data
        ws1 = wb.active
        ws1.title = "ATE test item check list"
        self._create_check_list_sheet(ws1, report)

        # Sheet 2: "Equipment Solution" - 设备方案（去重 + 选择列）
        ws2 = wb.create_sheet("Equipment Solution")
        self._create_equipment_solution_sheet(ws2, report)

        # Sheet 3: "vi_parameter_table"
        ws3 = wb.create_sheet("vi_parameter_table")
        self._create_vi_parameter_sheet(ws3, report)

        # Sheet 4: "Variables" - 所有变量
        ws4 = wb.create_sheet("Variables")
        self._create_variables_sheet(ws4, report)

        # Sheet 5: "Enum_Reference" - 枚举值常量参考（不影响序列文件生成）
        ws5 = wb.create_sheet("Enum_Reference")
        self._create_enum_reference_sheet(ws5, report)

        # Add bidirectional hyperlinks between Sheet 1 and Sheet 3
        self._add_step_hyperlinks(ws1, ws3)

        wb.save(output_path)

    def _create_check_list_sheet(self, ws, report: TestCaseReport):
        """Create 'ATE test item check list' sheet matching template with 4 header rows."""
        # Row 1: Title
        ws.merge_cells('A1:AD1')
        ws.cell(row=1, column=1, value="ATE test item check list")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Row 2: Product info
        ws.merge_cells('A2:T2')
        ws.cell(row=2, column=1, value="Product Type  FCT")
        ws.merge_cells('U2:AC2')
        ws.cell(row=2, column=21, value="EAU:                                                                 PCS/Y")
        ws.cell(row=2, column=30, value="Version:")

        # Row 3: ATE QTY info
        ws.merge_cells('A3:T3')
        ws.cell(row=3, column=1, value="ATE QTY :")
        ws.merge_cells('U3:AC3')
        ws.cell(row=3, column=21, value="Test time：                                                   S/Unit")
        ws.cell(row=3, column=30, value="")

        # Row 4: Column headers (30 columns)
        header_defs = [
            (1, "Step_No.\n序号", 'gray', 'center'),
            (2, "Step", 'gray', 'center'),
            (3, "Test project items\n测试项目", 'gray', 'left'),
            (4, "Describition of Test Requirments\n测试要求描述", 'gray', 'left'),
            (5, "step type\n步骤类型", 'gray', 'center'),
            (6, "instrument_vi\n仪器库vi", 'gray', 'center'),
            (7, "limits\n限值", 'gray', 'left'),
            (8, "USL\t上限", 'gray', 'center'),
            (9, "LSL\t下限", 'gray', 'center'),
            (10, "unit\n单位", 'gray', 'center'),
            (11, "format\n格式", 'gray', 'center'),
            (12, "run mode\n运行模式", 'gray', 'center'),
            (13, "wait(s)\n等待时间", 'gray', 'center'),
            (14, "adapter\n适配器", 'gray', 'center'),
            (15, "additional results\n附加结果", 'gray', 'left'),
            (16, "settings\n设置", 'gray', 'left'),
            (17, "Comment\n备注", 'gray', 'left'),
            (18, "Input Signals", 'yellow', 'left'),
            (19, "Output Loads", 'yellow', 'left'),
            (20, "Precondition Item/Status", 'yellow', 'left'),
            (21, "Command\n控制命令", 'gray', 'center'),
            (22, "Test point\t测试点", 'gray', 'center'),
            (23, "Measurement Value\t测量值", 'gray', 'center'),
            (24, "精度 (%)\n精度", 'gray', 'center'),
            (25, "Method Check\n检验方法", 'gray', 'center'),
            (26, "Equipment Solution\n设备方案", 'gray', 'center'),
            (27, "Parameter\n参数", 'gray', 'center'),
            (28, "Logfile confirm\t记录确认", 'gray', 'center'),
            (29, "TE comments", 'gray', 'center'),
            (30, "20230419 备注", 'gray', 'center'),
        ]

        for col_idx, header, fill_type, align in header_defs:
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_yellow if fill_type == 'yellow' else self.header_fill_gray
            cell.border = self.border
            cell.alignment = self.center_align if align == 'center' else self.left_align

        # Data rows start from row 5
        for tc in report.test_cases:
            row_num = ws.max_row + 1
            row_data = [
                tc.step_no,           # Step_N0
                tc.step,              # Step
                tc.test_project,      # Test project items
                tc.description,       # Describition of Test Requirments
                tc.step_type,         # step type
                tc.instrument_vi,     # instrument_vi (VI filename)
                tc.limits,            # limits
                tc.usl,               # USL
                tc.lsl,               # LSL
                tc.unit,              # unit
                tc.format,            # format
                tc.run_mode,          # run mode
                tc.wait_seconds,         # wait(s)
                tc.adapter,           # adapter
                tc.additional_results, # additional results
                tc.settings,          # settings
                tc.comments,          # comments - step.Comment
                tc.input_signals,     # Input Signals
                tc.output_loads,      # Output Loads
                tc.precondition,      # Precondition Item/Status
                "",                   # Command (manual entry)
                tc.test_point,        # Test point
                tc.measurement_value, # Measurement Value
                tc.precision,         # 精度(%)
                tc.method_check,      # Method Check
                tc.equipment,         # Equipment Solution
                tc.parameter,         # Parameter
                tc.logfile_confirm,   # Logfile confirm
                tc.te_comments,       # TE comments
                tc.remarks            # 备注
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = self._write_cell_safe(ws, row_num, col_idx, value)
                cell.border = self.border
                # Data row alignment: Step_N0 centered, rest left
                if col_idx == 1:
                    cell.alignment = self.center_align
                else:
                    cell.alignment = self.left_align

        self._auto_width_columns(ws)

    def _create_equipment_solution_sheet(self, ws, report: TestCaseReport) -> None:
        """Create 'Equipment Solution' sheet with unique equipment and selection columns.

        列结构: Equipment Solution | 推荐1 | 推荐2 | 推荐3 | comfirm
        从所有测试项的 equipment 字段中按 ", " 分隔并去重。
        """
        # 提取所有 equipment 字段（按 ", " 分割）并去重，保留首次出现顺序
        seen = set()
        unique_equipments = []
        for tc in report.test_cases:
            if not tc.equipment:
                continue
            for item in tc.equipment.split(","):
                name = item.strip()
                if not name or name in seen:
                    continue
                seen.add(name)
                unique_equipments.append(name)

        # 仅保留 20 个标准仪器类别，其它过滤掉
        allowed = {
            "ACSource", "Audio", "BatterySimulation", "DCSource", "DMM",
            "ElectronicLoad", "IOControl", "LEDAnalyzer", "MeasureModule",
            "Motor", "Optical", "Power", "Protocal", "RF", "Safety",
            "Scope", "SignalAcquisition", "SignalGeneration",
            "TemperatureHumidity", "WIFIBLE",
        }
        # 按 allowed 顺序输出（保持稳定顺序）
        unique_equipments = [e for e in unique_equipments if e in allowed]

        # Header rows（仿照参考模板的格式）
        # Row 1: 标题
        ws.cell(row=1, column=1, value="Equipment Solution")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Row 2: Model 行（模板中第 2 行的占位）
        ws.cell(row=2, column=2, value="Model")

        # Row 3: 列头
        headers = [
            "Equipment Solution\n设备方案",
            "推荐1",
            "推荐2",
            "推荐3",
            "comfirm",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_gray
            cell.border = self.border
            cell.alignment = self.center_align

        # Data rows
        for idx, equip in enumerate(unique_equipments, start=4):
            cell = ws.cell(row=idx, column=1, value=equip)
            cell.border = self.border
            cell.alignment = self.left_align
            # 推荐1/2/3/comfirm 留空供填写
            for col_idx in range(2, 6):
                empty_cell = ws.cell(row=idx, column=col_idx, value="")
                empty_cell.border = self.border
                empty_cell.alignment = self.left_align

        # 调整列宽
        ws.column_dimensions['A'].width = 40
        for col in ('B', 'C', 'D', 'E'):
            ws.column_dimensions[col].width = 20

    def _add_step_hyperlinks(self, ws1, ws2) -> None:
        """Add bidirectional hyperlinks between Sheet1 Step_No and Sheet2 Step_No."""
        # Build Sheet2 Step_No → row mapping (skip VI definition rows)
        s2_map = {}  # {step_no: row}
        for r in range(2, ws2.max_row + 1):
            step_no = ws2.cell(row=r, column=1).value
            if step_no:
                s2_map[str(step_no).strip()] = r

        # Build Sheet1 Step_No → row mapping (skip header rows)
        s1_map = {}  # {step_no: row}
        for r in range(5, ws1.max_row + 1):
            step_no = ws1.cell(row=r, column=1).value
            if step_no:
                s1_map[str(step_no).strip()] = r

        # Add hyperlinks from Sheet1 → Sheet2
        for step_no, s2_row in s2_map.items():
            if step_no in s1_map:
                s1_row = s1_map[step_no]
                cell = ws1.cell(row=s1_row, column=1)
                cell.hyperlink = f"#vi_parameter_table!A{s2_row}"
                cell.font = Font(color="0563C1", underline="single", size=10)

        # Add hyperlinks from Sheet2 → Sheet1
        for step_no, s1_row in s1_map.items():
            if step_no in s2_map:
                s2_row = s2_map[step_no]
                cell = ws2.cell(row=s2_row, column=1)
                cell.hyperlink = f"#'ATE test item check list'!A{s1_row}"
                cell.font = Font(color="0563C1", underline="single", size=10)

    def _create_vi_parameter_sheet(self, ws, report: TestCaseReport):
        """Create 'vi_parameter_table' sheet matching template format.

        Structure: Each VI has one definition row, followed by all steps that use it.
        - VI definition row: Step_No empty, instrument_vi_name has value, parameters are parameter names
        - Step rows: Step_No has value, instrument_vi_name empty, parameters are parameter values

        For parameters with enum values, adds dropdown validation using helper columns
        to avoid Excel's 255-character formula limit.
        """
        # Header row
        headers = [
            "Step_No", "instrument_vi_name",
            "Parameter1", "Parameter2", "Parameter3", "Parameter4", "Parameter5",
            "Parameter6", "Parameter7", "Parameter8", "Parameter9", "Parameter10"
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_gray
            cell.alignment = self.center_align
            cell.border = self.border

        # Group test cases by VI
        vi_groups = {}  # {vi_name: [tc1, tc2, ...]}
        for tc in report.test_cases:
            if tc.instrument_vi:
                if tc.instrument_vi not in vi_groups:
                    vi_groups[tc.instrument_vi] = []
                vi_groups[tc.instrument_vi].append(tc)

        # Font colors for parameter direction
        font_green = Font(color="008000", size=10)  # input params
        font_red = Font(color="FF0000", size=10)    # output params

        # Write data: VI definition row + step rows for each VI
        row_num = 1
        for vi_name, tcs in vi_groups.items():
            # Get all parameter names from first test case (input + output)
            first_tc = tcs[0]
            all_params = list(first_tc.module_parameters)
            param_names = [mp.name for mp in all_params]
            param_directions = [mp.direction for mp in all_params]  # "0"=input, "1"=output

            # VI definition row: Step_No empty, instrument_vi_name has value, parameters are names
            row_num += 1
            vi_data_start_row = row_num
            values = ["", vi_name]
            for pname in param_names:
                values.append(pname)
            while len(values) < 12:
                values.append("")
            values = values[:12]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = self.border
                cell.alignment = self.left_align
                # Color parameter name font by direction
                param_idx = col_idx - 3  # columns 3+ are parameters
                if 0 <= param_idx < len(param_directions):
                    if param_directions[param_idx] == "0":
                        cell.font = font_green
                    elif param_directions[param_idx] == "1":
                        cell.font = font_red

            # Step rows for this VI - show all parameter values
            for tc in tcs:
                row_num += 1
                tc_all_params = list(tc.module_parameters)
                param_values = []
                for i, mp in enumerate(tc_all_params):
                    val = mp.value_expr or mp.default_value
                    # If param has enum values and value is a numeric index, convert to enum name
                    if mp.enum_values and val:
                        try:
                            idx = int(float(val))
                            if 0 <= idx < len(mp.enum_values):
                                val = mp.enum_values[idx]
                        except (ValueError, TypeError):
                            pass
                    param_values.append(val)

                values = [tc.step_no, ""]  # Step_No has value, instrument_vi_name empty
                for pv in param_values:
                    values.append(pv)
                while len(values) < 12:
                    values.append("")
                values = values[:12]

                for col_idx, value in enumerate(values, 1):
                    cell = self._write_cell_safe(ws, row_num, col_idx, value)
                    cell.border = self.border
                    cell.alignment = self.left_align

        self._auto_width_columns(ws)

    def _auto_width_columns(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_variables_sheet(self, ws, report: TestCaseReport):
        """Create 'Variables' sheet with all variables from the sequence file."""
        headers = ["Container", "Variable Name", "Value", "Type"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_gray
            cell.alignment = self.center_align
            cell.border = self.border

        # Collect all unique container names, sorted logically
        all_containers = sorted(set(v.container for v in report.variables),
                                key=lambda c: (0 if c == "FileGlobals" else 1, c))
        row_num = 1
        for container in all_containers:
            vars_in_container = [v for v in report.variables if v.container == container]
            for v in vars_in_container:
                row_num += 1
                for col_idx, value in enumerate([v.container, v.name, v.value, v.type], 1):
                    cell = self._write_cell_safe(ws, row_num, col_idx, value)
                    cell.border = self.border
                    cell.alignment = self.left_align

        self._auto_width_columns(ws)

    def _create_enum_reference_sheet(self, ws, report: TestCaseReport):
        """Create 'Enum_Reference' sheet with all VI enum constant values.

        结构（与 ATETestPlan 模板一致）:
            Row 1: 标题
            Row 2: 描述
            Row 3: 空
            Row 4: 列头 (VI | 参数名 | 索引 | 常量名称)
            Row 5+: 扁平数据行

        此表仅供参考，不影响序列文件生成。
        """
        # Row 1: 标题
        ws.cell(row=1, column=1, value="VI 枚举常量值参考")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Row 2: 描述
        ws.cell(row=2, column=1, value="此表仅供参考，下拉列表值可从此处复制到 Parameter 列使用")
        ws.cell(row=2, column=1).font = Font(italic=True, size=10)

        # Row 4: 列头
        headers = ["VI", "参数名", "索引", "常量名称"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill_gray
            cell.alignment = self.center_align
            cell.border = self.border

        # 收集所有 (vi_name, param_name) -> [enum_values]
        vi_enums = {}
        for tc in report.test_cases:
            if not tc.instrument_vi:
                continue
            for mp in tc.module_parameters:
                if mp.enum_values:
                    key = (tc.instrument_vi, mp.name)
                    if key not in vi_enums:
                        vi_enums[key] = mp.enum_values

        # 扁平写入: 每个 (VI, 参数) 一段，枚举值逐行展开
        row_num = 5
        for (vi_name, param_name), enum_values in vi_enums.items():
            for idx, enum_val in enumerate(enum_values):
                values = [vi_name, param_name, idx, enum_val]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.border = self.border
                    if col_idx in (3,):
                        cell.alignment = self.center_align
                    else:
                        cell.alignment = self.left_align
                row_num += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 50