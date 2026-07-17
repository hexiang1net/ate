"""文档读取器，支持多种格式和图片提取。"""
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import List

logger = logging.getLogger(__name__)

# 图片最小尺寸（像素），小于此值的图片通常是 logo/图标，跳过
MIN_IMAGE_SIZE = 100


@dataclass
class PageImage:
    """从文档中提取的图片。"""
    page_num: int        # 页码（PDF）或段落位置（DOCX）
    image_bytes: bytes   # 图片原始字节
    media_type: str      # "image/png", "image/jpeg" 等
    caption: str = ""    # 图片附近的文字说明（如有）


@dataclass
class DocumentContent:
    """文档内容，包含文本和图片。"""
    text: str
    images: List[PageImage] = field(default_factory=list)
    source_format: str = ""   # "pdf", "docx", "xlsx", "txt" 等


def read_document(path: str, extract_images: bool = True) -> DocumentContent:
    """读取文档内容，返回文本和图片。

    支持格式: .docx, .doc, .xlsx, .xls, .pdf, .html, .htm, .md, .txt

    Args:
        path: 文档路径
        extract_images: 是否提取图片（禁用可节省处理时间）
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    ext = p.suffix.lower()
    readers = {
        ".docx": _read_docx,
        ".doc": _read_doc,
        ".xlsx": _read_xlsx,
        ".xls": _read_xls,
        ".pdf": _read_pdf,
        ".html": _read_html,
        ".htm": _read_html,
        ".md": _read_text,
        ".txt": _read_text,
    }

    reader = readers.get(ext)
    if reader is None:
        raise ValueError(f"不支持的文件格式: {ext}，支持: {', '.join(readers.keys())}")

    return reader(path, extract_images)


def _read_docx(path: str, extract_images: bool = True) -> DocumentContent:
    """读取 Word 文档。"""
    try:
        from docx import Document
    except ImportError:
        raise ImportError("读取 Word 文档需要安装 python-docx: pip install python-docx")

    try:
        doc = Document(path)
    except Exception as e:
        err_msg = str(e)
        if "OLE2" in err_msg or "OOXML" in err_msg:
            raise ValueError(
                f"文件不是有效的 .docx 格式: {path}\n"
                "可能原因: 文件实际是 .doc 格式（旧版 Word），需要先另存为 .docx；"
                "或文件已损坏。"
            ) from e
        raise ValueError(f"无法读取 Word 文档: {path} — {err_msg}") from e

    parts = []
    images = []
    position = 0

    # 按文档顺序遍历段落和表格
    for element in doc.element.body:
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag
        if tag == "p":
            from docx.text.paragraph import Paragraph
            para = Paragraph(element, doc)
            text = para.text.strip()
            if text:
                parts.append(text)
            position += 1
        elif tag == "tbl":
            table_text = _parse_docx_table(element, doc)
            if table_text:
                parts.append(table_text)
            position += 1

    # 提取图片
    if extract_images:
        images = _extract_docx_images(doc, position)

    return DocumentContent(
        text="\n\n".join(parts),
        images=images,
        source_format="docx"
    )


def _extract_docx_images(doc, max_position: int) -> List[PageImage]:
    """从 DOCX 文档提取内嵌图片。"""
    images = []

    try:
        for i, shape in enumerate(doc.inline_shapes):
            try:
                image = shape.image
                image_bytes = image.blob
                content_type = image.content_type

                # 跳过太小的图片（可能是图标）
                # python-docx 无法直接获取图片尺寸，需要解析图片数据
                if len(image_bytes) < 1024:  # 小于 1KB 的图片可能是图标
                    logger.debug(f"跳过小图片: {len(image_bytes)} bytes")
                    continue

                # 确定 media_type
                media_type = content_type if content_type else "image/png"
                if not media_type.startswith("image/"):
                    media_type = "image/png"

                images.append(PageImage(
                    page_num=max_position,  # DOCX 没有页码概念，使用位置
                    image_bytes=image_bytes,
                    media_type=media_type,
                ))
            except Exception as e:
                logger.debug(f"提取 DOCX 图片 {i} 失败: {e}")
    except Exception as e:
        logger.debug(f"DOCX 图片提取失败: {e}")

    return images


def _parse_docx_table(tbl_element, doc) -> str:
    """解析 Word 表格为文本。"""
    from docx.table import Table
    table = Table(tbl_element, doc)
    rows = []
    for row in table.rows:
        cells = [cell.text.strip() for cell in row.cells]
        rows.append(" | ".join(cells))
    return "\n".join(rows)


def _read_xlsx(path: str, extract_images: bool = True) -> DocumentContent:
    """读取 Excel 文件。"""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        err_msg = str(e)
        if "OLE2" in err_msg or "OOXML" in err_msg:
            raise ValueError(
                f"文件不是有效的 .xlsx 格式: {path}\n"
                "可能原因: 文件实际是 .xls 格式（旧版 Excel），需要先另存为 .xlsx；"
                "或文件已损坏。"
            ) from e
        raise ValueError(f"无法读取 Excel 文件: {path} — {err_msg}") from e
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"[Sheet: {sheet_name}]")
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        parts.append("\n".join(rows))

    wb.close()
    return DocumentContent(
        text="\n\n".join(parts),
        images=[],  # Excel 图片提取较复杂，暂不支持
        source_format="xlsx"
    )


def _read_xls(path: str, extract_images: bool = True) -> DocumentContent:
    """读取旧版 Excel 文件。"""
    try:
        import xlrd
    except ImportError:
        raise ImportError("读取 .xls 文件需要安装 xlrd: pip install xlrd")

    wb = xlrd.open_workbook(path)
    parts = []

    for sheet in wb.sheets():
        parts.append(f"[Sheet: {sheet.name}]")
        rows = []
        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        parts.append("\n".join(rows))

    return DocumentContent(
        text="\n\n".join(parts),
        images=[],
        source_format="xls"
    )


def _read_pdf(path: str, extract_images: bool = True) -> DocumentContent:
    """读取 PDF 文档。"""
    parts = []
    images = []

    # 优先使用 PyMuPDF（更可靠）
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(path)

        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text("text")
            if text and text.strip():
                parts.append(f"[Page {page_num + 1}]\n{text.strip()}")

        if extract_images:
            images = _extract_pdf_images_from_doc(doc)

        doc.close()
        return DocumentContent(
            text="\n\n".join(parts),
            images=images,
            source_format="pdf"
        )
    except ImportError:
        pass

    # 备选：使用 pdfplumber
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("读取 PDF 文档需要安装 PyMuPDF 或 pdfplumber: pip install PyMuPDF pdfplumber")

    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text and text.strip():
                parts.append(f"[Page {i + 1}]\n{text.strip()}")

            tables = page.extract_tables()
            for table in tables:
                rows = []
                for row in table:
                    cells = [str(c) if c is not None else "" for c in row]
                    rows.append(" | ".join(cells))
                if rows:
                    parts.append("\n".join(rows))

        # 提取图片
        if extract_images:
            images = _extract_pdf_images(path)

    return DocumentContent(
        text="\n\n".join(parts),
        images=images,
        source_format="pdf"
    )


def _extract_pdf_images(path: str) -> List[PageImage]:
    """从 PDF 提取图片（使用 PyMuPDF）。"""
    images = []

    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.warning("PyMuPDF 未安装，无法提取 PDF 图片。安装命令: pip install PyMuPDF")
        return images

    try:
        doc = fitz.open(path)
        images = _extract_pdf_images_from_doc(doc)
        doc.close()
    except Exception as e:
        logger.warning(f"PDF 图片提取失败: {e}")

    return images


def _extract_pdf_images_from_doc(doc) -> List[PageImage]:
    """从已打开的 PyMuPDF 文档提取图片。"""
    images = []

    for page_num in range(len(doc)):
        page = doc[page_num]
        image_list = page.get_images(full=True)

        for img_index, img_info in enumerate(image_list):
            try:
                xref = img_info[0]
                base_image = doc.extract_image(xref)

                if base_image is None:
                    continue

                image_bytes = base_image["image"]
                width = base_image.get("width", 0)
                height = base_image.get("height", 0)

                # 跳过太小的图片
                if width < MIN_IMAGE_SIZE or height < MIN_IMAGE_SIZE:
                    logger.debug(f"跳过小图片: {width}x{height}")
                    continue

                # 确定 media_type
                ext = base_image.get("ext", "png")
                media_type_map = {
                    "png": "image/png",
                    "jpeg": "image/jpeg",
                    "jpg": "image/jpeg",
                    "bmp": "image/bmp",
                    "tiff": "image/tiff",
                }
                media_type = media_type_map.get(ext, "image/png")

                images.append(PageImage(
                    page_num=page_num + 1,  # 页码从 1 开始
                    image_bytes=image_bytes,
                    media_type=media_type,
                ))

            except Exception as e:
                logger.debug(f"提取 PDF 图片失败 (page {page_num + 1}, img {img_index}): {e}")

    return images


def _read_text(path: str, extract_images: bool = True) -> DocumentContent:
    """读取纯文本/Markdown 文件。"""
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    return DocumentContent(
        text=text,
        images=[],
        source_format=Path(path).suffix.lstrip(".")
    )


def _read_doc(path: str, extract_images: bool = True) -> DocumentContent:
    """读取旧版 Word .doc 文件（通过 Word COM）。"""
    try:
        import win32com.client
    except ImportError:
        raise ImportError(
            "无法直接读取 .doc 文件，因为缺少 pywin32 或不在 Windows 平台。\n"
            "请先将 .doc 转换为 .docx（用 Microsoft Word 另存为 .docx 格式），"
            "然后用转换后的 .docx 文件重新生成。"
        )

    word = None
    doc = None
    try:
        try:
            word = win32com.client.Dispatch("Word.Application")
        except Exception:
            raise RuntimeError(
                "无法启动 Microsoft Word，可能是未安装 Word 或 COM 组件不可用。\n"
                "请先将 .doc 转换为 .docx（用 Word / WPS / LibreOffice 另存为 .docx 格式），"
                "然后用转换后的 .docx 文件重新生成。"
            )
        word.Visible = False
        doc = word.Documents.Open(path)

        parts = []

        # 提取正文段落
        for para in doc.Paragraphs:
            text = para.Range.Text
            if text:
                # COM 返回的文本以 \r 结尾，strip 掉
                stripped = text.strip()
                if stripped:
                    parts.append(stripped)

        # 提取表格（合并单元格可能导致 Rows 迭代失败，回退到整体文本）
        for table in doc.Tables:
            try:
                rows = []
                for row in table.Rows:
                    cells = []
                    for cell in row.Cells:
                        cell_text = cell.Range.Text.strip()
                        cells.append(cell_text)
                    rows.append(" | ".join(cells))
                if rows:
                    parts.append("\n".join(rows))
            except Exception:
                # 合并单元格的表格无法逐行访问，用整体文本
                table_text = table.Range.Text.strip()
                if table_text:
                    parts.append(table_text)

        text = "\n\n".join(parts)
    finally:
        if doc is not None:
            try:
                doc.Close()
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass

    return DocumentContent(
        text=text,
        images=[],
        source_format="doc",
    )


def _read_html(path: str, extract_images: bool = True) -> DocumentContent:
    """读取 HTML 文件，提取文本内容。"""
    from html.parser import HTMLParser

    with open(path, "r", encoding="utf-8") as f:
        html_content = f.read()

    class HTMLTextExtractor(HTMLParser):
        """HTML 文本提取器，跳过脚本/样式，保留表格结构。"""

        def __init__(self):
            super().__init__()
            self.parts: list = []
            self._skip = False
            self._skip_tags = {"script", "style", "head", "noscript"}
            self._table_cells: list = []
            self._in_table = False
            self._current_row: list = []
            self._in_row = False
            self._cell_tag = ""
            self._block_tags = {"p", "div", "br", "h1", "h2", "h3", "h4", "h5", "h6",
                               "li", "tr", "hr", "pre", "blockquote", "section",
                               "article", "header", "footer", "main", "nav"}

        def handle_starttag(self, tag, attrs):
            tag_lower = tag.lower()
            if tag_lower in self._skip_tags:
                self._skip = True
            elif tag_lower == "table":
                self._in_table = True
            elif tag_lower == "tr" and self._in_table:
                self._in_row = True
                self._current_row = []
            elif tag_lower in ("td", "th") and self._in_row:
                self._cell_tag = tag_lower
            elif tag_lower in self._block_tags:
                if self._current_row:
                    self.parts.append("")  # 保持换行

        def handle_endtag(self, tag):
            tag_lower = tag.lower()
            if tag_lower in self._skip_tags:
                self._skip = False
            elif tag_lower == "table":
                self._in_table = False
            elif tag_lower == "tr" and self._in_row:
                self._in_row = False
                if self._current_row:
                    self._table_cells.append(" | ".join(self._current_row))
                    self._current_row = []
            elif tag_lower in ("td", "th"):
                self._cell_tag = ""

        def handle_data(self, data):
            if self._skip:
                return
            text = data.strip()
            if not text:
                return
            if self._cell_tag:
                self._current_row.append(text)
            else:
                self.parts.append(text)

        def get_text(self) -> str:
            # 先放表格（转文本），再放正文
            result = []
            if self._table_cells:
                result.append("\n".join(self._table_cells))
            if self.parts:
                result.append("\n\n".join(self.parts))
            return "\n\n".join(result)

    extractor = HTMLTextExtractor()
    extractor.feed(html_content)
    text = extractor.get_text()

    return DocumentContent(
        text=text,
        images=[],
        source_format="html",
    )
